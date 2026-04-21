import streamlit as st
import pytesseract
from pdf2image import convert_from_bytes
import pandas as pd
import re
import tempfile
import os
import time
import base64
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font

# =========================
# CONFIG
# =========================
st.set_page_config(page_title="THL PDF TO EXCEL", layout="wide")

# =========================
# SESSION
# =========================
if "processing" not in st.session_state:
    st.session_state.processing = False
if "done" not in st.session_state:
    st.session_state.done = False
if "clear_uploader" not in st.session_state:
    st.session_state.clear_uploader = False
if "last_uploaded_names" not in st.session_state:
    st.session_state.last_uploaded_names = []
if "excel_file" not in st.session_state:
    st.session_state.excel_file = None

# =========================
# STYLE (GIỮ NGUYÊN)
# =========================
st.markdown("""
<style>
header, #MainMenu, footer {visibility: hidden;}
.block-container {padding-top: 0.5rem !important;}
.stApp { background: #f1f5f9; }

.header {
    font-size:22px;
    font-weight:700;
    margin-bottom:10px;
}

[data-testid="stFileUploader"] {
    border: 2px dashed #93c5fd;
    padding: 25px;
    border-radius: 18px;
    background: white;
    transition: 0.3s;
}
[data-testid="stFileUploader"]:hover {
    border-color:#3b82f6;
}

div.stButton > button {
    background: linear-gradient(135deg,#3b82f6,#22c55e);
    color:white;
    border:none;
    border-radius:12px;
    padding:12px 24px;
    font-weight:600;
    font-size:15px;
    box-shadow:0 4px 14px rgba(0,0,0,0.15);
    transition: all 0.25s ease;
}
div.stButton > button:hover {
    transform: translateY(-2px) scale(1.02);
}

.new-btn button {
    background: linear-gradient(135deg,#f59e0b,#ef4444) !important;
}

.process-btn {
    margin-top: 25px;
    margin-bottom: 15px;
}

.file-row {
    margin-top:12px;
    padding:10px;
    border-radius:12px;
    background:white;
    box-shadow:0 2px 8px rgba(0,0,0,0.05);
}

.progress {
    height:8px;
    background:#e5e7eb;
    border-radius:999px;
    overflow:hidden;
    margin-top:6px;
}
.progress-bar {
    height:100%;
    background:linear-gradient(90deg,#3b82f6,#22c55e);
}

.loading {
    font-size:14px;
    color:#475569;
    margin-top:10px;
}
</style>
""", unsafe_allow_html=True)

# =========================
# HEADER
# =========================
st.markdown('<div class="header">🚀 THL PDF → EXCEL </div>', unsafe_allow_html=True)

# =========================
# UPLOADER
# =========================
uploader_key = "uploader_1" if not st.session_state.clear_uploader else "uploader_2"

uploaded_files = st.file_uploader(
    "📂 Chọn file PDF",
    type=["pdf"],
    accept_multiple_files=True,
    key=uploader_key
)

current_names = [f.name for f in uploaded_files] if uploaded_files else []

if current_names != st.session_state.last_uploaded_names:
    st.session_state.processing = False
    st.session_state.done = False
    st.session_state.last_uploaded_names = current_names

# =========================
# OCR (FINAL FIX)
# =========================
def ocr_extract(img):

    def normalize(text):
        text = text.upper()
        text = text.replace("P8", "PR")
        text = text.replace("S0", "SO")
        return text

    def is_valid_header(text):
        t = text.upper()
        return "TIEN PHONG" in t or "NHUATIENPHONG" in t

    def extract_from_text(text):

        if not is_valid_header(text):
            return None, None, None

        sm = None
        prso = None
        date = None

        for line in text.split("\n"):
            raw = line.strip()
            clean = normalize(raw)

            # SM
            if not sm:
                m = re.search(r"(SM\d{4}\.\d{4})", clean)
                if m:
                    sm = m.group(1)

            # ===== FIX QUAN TRỌNG: LẤY NGUYÊN CỤM PR/SO =====
            if not prso:
                m = re.search(r"(PR\d{4}\.\d{4}\s*/\s*SO\d{4}\.\d{4})", clean)
                if m:
                    prso = m.group(1).replace(" ", "")

            # fallback OCR lỗi nhẹ
            if not prso:
                m = re.search(r"(P[R8]\d{4}\.\d{4}).*?(S[O0]\d{4}\.\d{4})", clean)
                if m:
                    pr = m.group(1).replace("P8", "PR")
                    so = m.group(2).replace("S0", "SO")
                    prso = f"{pr}/{so}"

            # PR riêng
            if not prso:
                m = re.search(r"(P[R8]|R)\d{4}\.\d{4}", clean)
                if m:
                    val = m.group(0)
                    if val.startswith("R"):
                        val = "P" + val
                    prso = val.replace("P8", "PR")

            # SO fallback
            if not prso:
                m = re.search(r"S[O0]\d{4}\.\d{4}", clean)
                if m:
                    prso = m.group(0).replace("S0", "SO")

            # DATE
            if not date:
                d = re.search(r"(\d{2}/\d{2}/\d{4})", raw)
                if d:
                    date = d.group(1)

        return sm, prso, date

    w, h = img.size

    header = img.crop((0, 0, w, int(h * 0.25)))
    quick = pytesseract.image_to_string(header, config='--oem 3 --psm 6')

    if not is_valid_header(quick):
        return None, None, None

    for variant in [
        img.crop((0, 0, w, int(h * 0.4))),
        img
    ]:
        text = pytesseract.image_to_string(variant, config='--oem 3 --psm 6')
        sm, prso, date = extract_from_text(text)

        if (sm or prso) and date:
            return sm, prso, date

    return None, None, None

# =========================
# PROCESS
# =========================
def extract_pdf(file, box):

    results = []
    images = convert_from_bytes(file.read(), dpi=150)
    total = len(images)

    for i, img in enumerate(images, 1):

        percent = int((i/total)*100)

        box.markdown(f"""
<div class="file-row">
📄 {file.name} — Trang {i}/{total}
<div class="progress">
<div class="progress-bar" style="width:{percent}%"></div>
</div>
</div>
""", unsafe_allow_html=True)

        sm, prso, date = ocr_extract(img)

        if sm or prso:
            results.append({
                "SM": sm or "",
                "PR/SO": prso or "",
                "Ngày": date or "",
                "Trang": i
            })

    return results

# =========================
# MAIN
# =========================
if uploaded_files:

    boxes = [st.empty() for _ in uploaded_files]

    if not st.session_state.processing and not st.session_state.done:
        st.markdown('<div class="process-btn">', unsafe_allow_html=True)
        if st.button("🚀 Bắt đầu xử lý"):
            st.session_state.processing = True
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    if st.session_state.processing:

        st.markdown('<div class="loading">⏳ Đang xử lý... vui lòng chờ</div>', unsafe_allow_html=True)

        tmp_excel = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")

        with pd.ExcelWriter(tmp_excel.name, engine='openpyxl') as writer:

            has_data = False

            for i, f in enumerate(uploaded_files):

                data = extract_pdf(f, boxes[i])

                if data:
                    has_data = True

                    df = pd.DataFrame(data)
                    df.insert(0, "STT", range(1, len(df)+1))

                    name = os.path.splitext(f.name)[0][:31]
                    df.to_excel(writer, sheet_name=name, index=False)

            if not has_data:
                df = pd.DataFrame([{"Thông báo": "Không có dữ liệu hợp lệ"}])
                df.to_excel(writer, sheet_name="KET_QUA", index=False)

        wb = load_workbook(tmp_excel.name)

        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = max(len(str(c.value)) if c.value else 0 for c in col)
                ws.column_dimensions[col[0].column_letter].width = max_len + 3

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border

            for cell in ws[1]:
                cell.font = Font(bold=True)

        wb.save(tmp_excel.name)

        st.session_state.excel_file = tmp_excel.name
        st.session_state.processing = False
        st.session_state.done = True
        st.rerun()

# =========================
# DOWNLOAD
# =========================
if st.session_state.done:

    st.success("🎉 HOÀN THÀNH !!!")

    with open(st.session_state.excel_file, "rb") as f:
        data = f.read()

    b64 = base64.b64encode(data).decode()

    st.markdown(f"""
        <iframe src="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" style="display:none;"></iframe>
    """, unsafe_allow_html=True)

    st.markdown('<div class="new-btn">', unsafe_allow_html=True)
    if st.button("🔄 XỬ LÝ FILE MỚI"):
        st.session_state.done = False
        st.session_state.clear_uploader = not st.session_state.clear_uploader
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)
